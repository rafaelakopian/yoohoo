"""Generic file parser for Excel (.xlsx) and CSV files."""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from typing import Any

import structlog
from openpyxl import load_workbook

logger = structlog.get_logger()

MAX_ROWS = 1000


@dataclass
class ParseResult:
    """Result of parsing an uploaded file."""

    headers: list[str]
    rows: list[dict[str, Any]]
    total_rows: int


def parse_file(file_data: bytes, file_type: str) -> ParseResult:
    """Parse an uploaded file and return headers + row data.

    Supports xlsx and csv. Auto-detects header row for xlsx (row with most
    non-empty cells in rows 1-5). For CSV, auto-detects encoding and delimiter.

    Raises ValueError if file has >MAX_ROWS data rows.
    """
    if file_type == "csv":
        return _parse_csv(file_data)
    return _parse_xlsx(file_data)


def _parse_xlsx(file_data: bytes) -> ParseResult:
    """Parse Excel file with auto-detected header row."""
    wb = load_workbook(filename=io.BytesIO(file_data), read_only=True, data_only=True)

    # Try to find a sheet with "leerling" in name, fall back to first
    sheet = None
    for name in wb.sheetnames:
        if "leerling" in name.lower():
            sheet = wb[name]
            break
    if sheet is None:
        sheet = wb.worksheets[0]

    all_rows = list(sheet.iter_rows(min_row=1, values_only=True))
    wb.close()

    if len(all_rows) == 0:
        raise ValueError("Bestand is leeg.")

    if len(all_rows) == 1:
        # Only header row, no data
        headers = [str(h).strip() if h else "" for h in all_rows[0]]
        return ParseResult(headers=[h for h in headers if h], rows=[], total_rows=0)

    # Auto-detect header row: scan rows 0-4, pick row with most non-empty cells
    scan_limit = min(5, len(all_rows))
    best_row_idx = 0
    best_count = 0
    for i in range(scan_limit):
        count = sum(1 for c in all_rows[i] if c is not None and str(c).strip())
        if count > best_count:
            best_count = count
            best_row_idx = i

    header_row = all_rows[best_row_idx]
    headers = [str(h).strip() if h else "" for h in header_row]

    # Data rows start after header
    data_start = best_row_idx + 1
    data_rows_raw = all_rows[data_start:]

    # Filter out completely empty rows
    data_rows = []
    for row in data_rows_raw:
        if row and any(cell is not None and str(cell).strip() for cell in row):
            data_rows.append(row)

    total_rows = len(data_rows)
    if total_rows > MAX_ROWS:
        raise ValueError(f"Bestand bevat meer dan {MAX_ROWS} rijen ({total_rows} gevonden).")

    # Convert to list of dicts keyed by header name
    rows: list[dict[str, Any]] = []
    for row in data_rows:
        row_dict: dict[str, Any] = {}
        for col_idx, header in enumerate(headers):
            if not header:
                continue
            value = row[col_idx] if col_idx < len(row) else None
            if value is not None:
                value = str(value).strip()
                if not value:
                    value = None
            row_dict[header] = value
        rows.append(row_dict)

    logger.info("file_parser.xlsx_parsed", headers=len(headers), rows=total_rows)
    return ParseResult(headers=[h for h in headers if h], rows=rows, total_rows=total_rows)


def _parse_csv(file_data: bytes) -> ParseResult:
    """Parse CSV file with encoding and delimiter auto-detection."""
    # Try encodings in order: UTF-8 BOM, UTF-8, Latin-1
    text = None
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = file_data.decode(encoding)
            break
        except (UnicodeDecodeError, ValueError):
            continue

    if text is None:
        raise ValueError("Kan het bestand niet lezen — onbekende tekencodering.")

    # Detect delimiter
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ","

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    all_rows = list(reader)

    if len(all_rows) == 0:
        raise ValueError("Bestand is leeg.")

    headers = [h.strip() for h in all_rows[0]]

    if len(all_rows) == 1:
        return ParseResult(headers=[h for h in headers if h], rows=[], total_rows=0)

    data_rows = [r for r in all_rows[1:] if any(c.strip() for c in r)]

    total_rows = len(data_rows)
    if total_rows > MAX_ROWS:
        raise ValueError(f"Bestand bevat meer dan {MAX_ROWS} rijen ({total_rows} gevonden).")

    rows: list[dict[str, Any]] = []
    for row in data_rows:
        row_dict: dict[str, Any] = {}
        for col_idx, header in enumerate(headers):
            if not header:
                continue
            value = row[col_idx].strip() if col_idx < len(row) else None
            if not value:
                value = None
            row_dict[header] = value
        rows.append(row_dict)

    logger.info("file_parser.csv_parsed", headers=len(headers), rows=total_rows)
    return ParseResult(headers=[h for h in headers if h], rows=rows, total_rows=total_rows)
