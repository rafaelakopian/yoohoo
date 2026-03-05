from io import BytesIO

import structlog
from openpyxl import load_workbook

from app.modules.products.school.student.schemas import StudentCreate

logger = structlog.get_logger()

# Column mapping: Excel header → StudentCreate field
COLUMN_MAP = {
    "voornaam student": "first_name",
    "achternaam student": "last_name",
    "lesdag": "lesson_day",
    "lesduur": "lesson_duration",
    "lestijd": "lesson_time",
    "ouder of voogd": "guardian_name",
    "relatie ouder/voogd": "guardian_relationship",
    "telefoonnummer privé/werk ouder/voogd": "guardian_phone",
}


def parse_excel(file_data: bytes) -> tuple[list[StudentCreate], list[str]]:
    """Parse an Excel file with student data from the Leerlingenlijst sheet.

    Returns a tuple of (parsed students, error messages).
    """
    wb = load_workbook(filename=BytesIO(file_data), read_only=True, data_only=True)

    # Try to find the Leerlingenlijst sheet
    sheet = None
    for name in wb.sheetnames:
        if "leerling" in name.lower():
            sheet = wb[name]
            break

    if sheet is None:
        # Fall back to first sheet
        sheet = wb.worksheets[0]
        logger.warning("excel_import.sheet_fallback", sheet_name=sheet.title)

    # Read header row (row 3, 0-indexed row 2)
    rows = list(sheet.iter_rows(min_row=1, values_only=True))
    if len(rows) < 4:
        return [], ["Excel file has too few rows (expected headers at row 3, data from row 4)"]

    header_row = rows[2]  # Row 3 (0-indexed)
    headers = [str(h).strip().lower() if h else "" for h in header_row]

    # Map column indices to fields
    col_mapping: dict[int, str] = {}
    for col_idx, header in enumerate(headers):
        for excel_header, field_name in COLUMN_MAP.items():
            if excel_header in header:
                col_mapping[col_idx] = field_name
                break

    if not col_mapping:
        return [], ["Could not find any recognized column headers in row 3"]

    students: list[StudentCreate] = []
    errors: list[str] = []

    for row_idx, row in enumerate(rows[3:], start=4):  # Data starts at row 4
        # Skip empty rows
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        row_data: dict = {}
        for col_idx, field_name in col_mapping.items():
            if col_idx < len(row) and row[col_idx] is not None:
                value = str(row[col_idx]).strip()
                if value:
                    if field_name == "lesson_duration":
                        try:
                            row_data[field_name] = int(float(value))
                        except (ValueError, TypeError):
                            errors.append(f"Row {row_idx}: invalid lesson duration '{value}'")
                    else:
                        row_data[field_name] = value

        # Skip rows without a first name
        if not row_data.get("first_name"):
            continue

        try:
            students.append(StudentCreate(**row_data))
        except Exception as e:
            errors.append(f"Row {row_idx}: {str(e)}")

    wb.close()

    logger.info("excel_import.parsed", students=len(students), errors=len(errors))
    return students, errors
